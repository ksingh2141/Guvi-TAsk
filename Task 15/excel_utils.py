from openpyxl import load_workbook
from datetime import datetime


class ExcelUtils:

    @staticmethod
    def get_row_count(file, sheet):
        wb = load_workbook(file)
        sh = wb[sheet]
        return sh.max_row

    @staticmethod
    def read_data(file, sheet, row, col):
        wb = load_workbook(file)
        sh = wb[sheet]
        return sh.cell(row=row, column=col).value

    @staticmethod
    def write_data(file, sheet, row, col, data):
        wb = load_workbook(file)
        sh = wb[sheet]
        sh.cell(row=row, column=col).value = data
        wb.save(file)

    @staticmethod
    def update_test_result(file, sheet, row, result):

        wb = load_workbook(file)
        sh = wb[sheet]

        now = datetime.now()

        sh.cell(row=row, column=4).value = now.strftime("%Y-%m-%d")
        sh.cell(row=row, column=5).value = now.strftime("%H:%M:%S")
        sh.cell(row=row, column=7).value = result

        wb.save(file)