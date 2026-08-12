import openpyxl


class ExcelReader:
    """
    Utility class for reading Excel test data.
    """

    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)

    def get_sheet(self, sheet_name):
        return self.workbook[sheet_name]

    def row_count(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        return sheet.max_row

    def column_count(self, sheet_name):
        sheet = self.get_sheet(sheet_name)
        return sheet.max_column

    def read_cell(self, sheet_name, row, column):
        sheet = self.get_sheet(sheet_name)
        return sheet.cell(row=row, column=column).value

    def write_cell(self, sheet_name, row, column, value):
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row, column=column).value = value

    def save(self, file_path=None):
        if file_path:
            self.workbook.save(file_path)
        else:
            self.workbook.save(self.workbook.filename)

    def get_login_data(self, sheet_name):
        """
        Returns:
        [
            (username,password,expected),
            ...
        ]
        """
        sheet = self.get_sheet(sheet_name)

        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data