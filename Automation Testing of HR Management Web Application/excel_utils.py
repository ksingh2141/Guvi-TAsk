from openpyxl import load_workbook


class ExcelUtils:

    def __init__(self, file):
        self.file = file
        self.workbook = load_workbook(file)
        self.sheet = self.workbook.active

    def row_count(self):
        return self.sheet.max_row

    def read_data(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def write_data(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value
        self.workbook.save(self.file)   #